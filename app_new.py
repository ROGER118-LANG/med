import streamlit as st
from keras.models import load_model
from keras.layers import DepthwiseConv2D
from keras.utils import custom_object_scope
from PIL import Image, ImageOps
import numpy as np
from scipy.ndimage import zoom
import plotly.graph_objects as go
import io
import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
import hashlib
import matplotlib.pyplot as plt
import seaborn as sns

# Configuração da página
st.set_page_config(page_title="Visualização 3D de Raio-X", layout="wide")
np.set_printoptions(suppress=True)

# Inicializar estado da sessão
if 'historico_paciente' not in st.session_state:
    st.session_state.historico_paciente = {}
if 'logado' not in st.session_state:
    st.session_state.logado = False
if 'nome_usuario' not in st.session_state:
    st.session_state.nome_usuario = None
if 'setores_usuario' not in st.session_state:
    st.session_state.setores_usuario = []

# Arquivo para armazenar informações de login
ARQUIVO_LOGIN = 'info_login.xlsx'

# Definição dos caminhos dos modelos e rótulos
caminhos_modelos = {
    "Pneumologia": {
        "Pneumonia": "pneumonia_model.h5",
        "Tuberculose": "tuberculose_model.h5",
        "Câncer de Pulmão": "cancer_pulmao_model.h5"
    },
    "Neurologia": {
        "Tumor Cerebral": "tumor_cerebral_model.h5"
    },
    "Ortopedia": {
        "Braço Fraturado": "fractured_arm_model.h5",
        "Ruptura do Tendão de Aquiles": "achilles_tendon_rupture_model.h5",
        "ACL": "acl_model.h5",
        "Entorse de Tornozelo": "ankle_sprain_model.h5",
        "Fratura de Calcâneo": "calcaneus_fracture_model.h5"
    }
}

caminhos_rotulos = {
    "Pneumologia": {
        "Pneumonia": "pneumonia_labels.txt",
        "Tuberculose": "tuberculose_labels.txt",
        "Câncer de Pulmão": "cancer_pulmao_labels.txt"
    },
    "Neurologia": {
        "Tumor Cerebral": "tumor_cerebral_labels.txt"
    },
    "Ortopedia": {
        "Braço Fraturado": "fractured_arm_labels.txt",
        "Ruptura do Tendão de Aquiles": "achilles_tendon_rupture_labels.txt",
        "ACL": "acl_labels.txt",
        "Entorse de Tornozelo": "ankle_sprain_labels.txt",
        "Fratura de Calcâneo": "calcaneus_fracture_labels.txt"
    }
}

def custom_depthwise_conv2d(*args, **kwargs):
    kwargs.pop('groups', None)
    return DepthwiseConv2D(*args, **kwargs)

def carregar_modelo_e_rotulos(caminho_modelo, caminho_rotulos):
    try:
        if not os.path.exists(caminho_modelo):
            raise FileNotFoundError(f"Arquivo de modelo não encontrado: {caminho_modelo}")
        if not os.path.exists(caminho_rotulos):
            raise FileNotFoundError(f"Arquivo de rótulos não encontrado: {caminho_rotulos}")
        
        with custom_object_scope({'DepthwiseConv2D': custom_depthwise_conv2d}):
            modelo = load_model(caminho_modelo, compile=False)
        
        with open(caminho_rotulos, "r") as f:
            nomes_classes = f.readlines()
        return modelo, nomes_classes
    except Exception as e:
        st.error(f"Erro ao carregar modelo e rótulos: {str(e)}")
        return None, None

def prever(modelo, dados, nomes_classes):
    try:
        previsao = modelo.predict(dados)
        indice = np.argmax(previsao)
        nome_classe = nomes_classes[indice]
        pontuacao_confianca = float(previsao[0][indice])
        return nome_classe.strip(), pontuacao_confianca
    except Exception as e:
        st.error(f"Erro durante a previsão: {str(e)}")
        return None, None

def preprocessar_imagem(arquivo_carregado):
    try:
        bytes_imagem = arquivo_carregado.getvalue()
        imagem = Image.open(io.BytesIO(bytes_imagem)).convert("RGB")
        tamanho = (224, 224)
        imagem = ImageOps.fit(imagem, tamanho, Image.Resampling.LANCZOS)
        array_imagem = np.asarray(imagem)
        array_imagem_normalizado = (array_imagem.astype(np.float32) / 127.5) - 1
        dados = np.ndarray(shape=(1, 224, 224, 3), dtype=np.float32)
        dados[0] = array_imagem_normalizado
        return dados
    except Exception as e:
        st.error(f"Erro ao pré-processar imagem: {str(e)}")
        return None

def classificar_exame(id_paciente, opcao_modelo, arquivo_carregado):
    if arquivo_carregado is not None:
        st.write(f"Opção de modelo selecionada: {opcao_modelo}")
        
        setor, modelo = opcao_modelo.split('_', 1)
        if setor not in caminhos_modelos or modelo not in caminhos_modelos[setor]:
            st.error(f"Opção de modelo '{opcao_modelo}' não encontrada nos modelos disponíveis.")
            return None
        
        try:
            modelo, nomes_classes = carregar_modelo_e_rotulos(caminhos_modelos[setor][modelo], caminhos_rotulos[setor][modelo])
            
            if modelo is not None and nomes_classes is not None:
                imagem_processada = preprocessar_imagem(arquivo_carregado)
                
                if imagem_processada is not None:
                    nome_classe, pontuacao_confianca = prever(modelo, imagem_processada, nomes_classes)
                    
                    if nome_classe is not None and pontuacao_confianca is not None:
                        resultado = {
                            'data': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            'modelo': opcao_modelo,
                            'classe': nome_classe,
                            'confianca': pontuacao_confianca
                        }
                        
                        if id_paciente not in st.session_state.historico_paciente:
                            st.session_state.historico_paciente[id_paciente] = []
                        st.session_state.historico_paciente[id_paciente].append(resultado)
                        
                        st.success("Exame classificado com sucesso!")
                        return resultado
                    else:
                        st.error("Ocorreu um erro durante a previsão. Por favor, tente novamente.")
                else:
                    st.error("Falha ao pré-processar a imagem. Por favor, tente uma imagem diferente.")
            else:
                st.error("Falha ao carregar o modelo e rótulos. Por favor, verifique os arquivos e tente novamente.")
        except Exception as e:
            st.error(f"Ocorreu um erro durante a classificação: {str(e)}")
    else:
        st.error("Por favor, faça o upload de uma imagem primeiro.")
    return None

def hash_senha(senha):
    return hashlib.sha256(senha.encode()).hexdigest()

def inicializar_arquivo_login():
    if not os.path.exists(ARQUIVO_LOGIN):
        wb = Workbook()
        ws = wb.active
        ws.append(['Nome de Usuário', 'Senha', 'Último Login', 'Data de Expiração', 'Função', 'Setores'])
        senha_admin = hash_senha('123')
        ws.append(['admin', senha_admin, '', '', 'admin', 'Pneumologia,Neurologia,Ortopedia'])
        wb.save(ARQUIVO_LOGIN)

def verificar_login(nome_usuario, senha):
    try:
        wb = load_workbook(ARQUIVO_LOGIN)
        ws = wb.active
        for linha in ws.iter_rows(min_row=2, values_only=True):
            if linha[0] == nome_usuario and linha[1] == hash_senha(senha):
                eh_admin = len(linha) > 4 and linha[4] == 'admin'
                
                if not eh_admin:
                    if len(linha) > 3 and linha[3]:
                        data_expiracao = linha[3]
                        if isinstance(data_expiracao, datetime) and datetime.now() > data_expiracao:
                            return False, "Conta expirada", []
                
                setores = linha[5].split(',') if len(linha) > 5 and linha[5] else []
                return True, "Sucesso", setores
        
        return False, "Credenciais inválidas", []
    except Exception as e:
        st.error(f"Ocorreu um erro ao verificar o login: {str(e)}")
        return False, "Falha na verificação do login", []

def atualizar_ultimo_login(nome_usuario):
    wb = load_workbook(ARQUIVO_LOGIN)
    ws = wb.active
    for linha in ws.iter_rows(min_row=2):
        if linha[0].value == nome_usuario:
            linha[2].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            break
    wb.save(ARQUIVO_LOGIN)

def pagina_login():
    st.title("Login")
    nome_usuario = st.text_input("Nome de Usuário")
    senha = st.text_input("Senha", type="password")
    if st.button("Entrar"):
        sucesso_login, mensagem, setores = verificar_login(nome_usuario, senha)
        if sucesso_login:
            st.session_state.logado = True
            st.session_state.nome_usuario = nome_usuario
            st.session_state.setores_usuario = setores
            atualizar_ultimo_login(nome_usuario)
            st.success("Login realizado com sucesso!")
        else:
            st.error(mensagem)

def visualizar_historico_paciente(id_paciente):
    if id_paciente in st.session_state.historico_paciente:
        historico = st.session_state.historico_paciente[id_paciente]
        df = pd.DataFrame(historico)
        st.dataframe(df)
        
        st.subheader("Visualização do Histórico de Exames do Paciente")
        fig, ax = plt.subplots(figsize=(10, 6))
        sns.scatterplot(data=df, x='data', y='confianca', hue='modelo', size='confianca', ax=ax)
        ax.set_title(f"Confiança dos Exames ao Longo do Tempo para o Paciente {id_paciente}")
        ax.set_xlabel("Data")
        ax.set_ylabel("Pontuação de Confiança")
        st.pyplot(fig)
    else:
        st.info("Nenhum histórico encontrado para este paciente.")

def comparar_pacientes():
    st.subheader("Comparar Pacientes")
    ids_pacientes = list(st.session_state.historico_paciente.keys())
    if len(ids_pacientes) < 2:
        st.warning("É necessário pelo menos dois pacientes com histórico para comparar.")
        return
    
    paciente1 = st.selectbox("Selecione o primeiro paciente", ids_pacientes)
    paciente2 = st.selectbox("Selecione o segundo paciente", [id for id in ids_pacientes if id != paciente1])
    
    if st.button("Comparar"):
        df1 = pd.DataFrame(st.session_state.historico_paciente[paciente1])
        df2 = pd.DataFrame(st.session_state.historico_paciente[paciente2])
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 6))
        
        sns.boxplot(data=df1, x='modelo', y='confianca', ax=ax1)
        ax1.set_title(f"Paciente {paciente1}")
        ax1.set_ylim(0, 1)
        
        sns.boxplot(data=df2, x='modelo', y='confianca', ax=ax2)
        ax2.set_title(f"Paciente {paciente2}")
        ax2.set_ylim(0, 1)
        
        st.pyplot(fig)

def gerenciar_usuarios():
    st.header("Gerenciamento de Usuários")
    
    try:
        wb = load_workbook(ARQUIVO_LOGIN)
        ws = wb.active
        dados_usuario = {linha[0]: linha for linha in ws.iter_rows(min_row=2, values_only=True)}
        
        dados_usuario_limpos = [
            (linha if len(linha) == 6 else linha + (None,) * (6 - len(linha))) 
            for linha in dados_usuario.values()
        ]
        
        df_usuario = pd.DataFrame(dados_usuario_limpos, columns=["Nome de Usuário", "Senha", "Último Login", "Data de Expiração", "Função", "Setores"])
        
        st.dataframe(df_usuario)

        st.subheader("Adicionar Usuário")
        novo_nome_usuario = st.text_input("Novo Nome de Usuário")
        nova_senha = st.text_input("Nova Senha", type="password")
        nova_funcao = st.selectbox("Função", ["usuário", "admin"])
        dias_validade = st.number_input("Validade da Conta (dias)", min_value=1, value=7, step=1)
        novos_setores = st.multiselect("Setores", ["Pneumologia", "Neurologia", "Ortopedia"])
        
        if st.button("Adicionar Usuário"):
            if novo_nome_usuario and nova_senha:
                senha_hash = hash_senha(nova_senha)
                data_expiracao = datetime.now() + timedelta(days=dias_validade) if nova_funcao != "admin" else None
                ws.append([novo_nome_usuario, senha_hash, "", data_expiracao, nova_funcao, ",".join(novos_setores)])
                wb.save(ARQUIVO_LOGIN)
                st.success("Usuário adicionado com sucesso!")
            else:
                st.error("Por favor, forneça nome de usuário e senha.")

        st.subheader("Editar Usuário")
        editar_nome_usuario = st.selectbox("Selecione o Usuário para Editar", list(dados_usuario.keys()))
        senha_editada = st.text_input("Nova Senha para o Usuário Selecionado", type="password")
        funcao_editada = st.selectbox("Nova Função", ["usuário", "admin"])
        validade_editada = st.number_input("Nova Validade da Conta (dias)", min_value=1, value=7, step=1)
        setores_editados = st.multiselect("Novos Setores", ["Pneumologia", "Neurologia", "Ortopedia"])
        
        if st.button("Editar Usuário"):
            if senha_editada:
                senha_hash = hash_senha(senha_editada)
                for linha in ws.iter_rows(min_row=2):
                    if linha[0].value == editar_nome_usuario:
                        linha[1].value = senha_hash
                        linha[3].value = datetime.now() + timedelta(days=validade_editada) if funcao_editada != "admin" else None
                        linha[4].value = funcao_editada
                        linha[5].value = ",".join(setores_editados)
                        break
                wb.save(ARQUIVO_LOGIN)
                st.success("Usuário editado com sucesso!")
            else:
                st.error("Por favor, forneça uma nova senha.")

        st.subheader("Remover Usuário")
        remover_nome_usuario = st.selectbox("Selecione o Usuário para Remover", list(dados_usuario.keys()))
        if st.button("Remover Usuário"):
            ws.delete_rows(list(dados_usuario.keys()).index(remover_nome_usuario) + 2)
            wb.save(ARQUIVO_LOGIN)
            st.success("Usuário removido com sucesso!")
    
    except Exception as e:
        st.error(f"Ocorreu um erro durante o gerenciamento de usuários: {str(e)}")

def reduzir_resolucao_matriz(matriz_3d, fator=0.5):
    """Reduz a resolução da matriz 3D em um fator especificado."""
    matriz_reduzida = zoom(matriz_3d, (fator, fator, fator))
    return matriz_reduzida

def converter_raio_x_para_3d(imagem, profundidade=20):
    """Converte a imagem de Raio-X para uma matriz 3D."""
    # Converte a imagem para escala de cinza
    imagem_cinza = imagem.convert('L')
    # Converte para array numpy
    array_imagem = np.array(imagem_cinza)
    
    # Normaliza valores para o intervalo [0, 1]
    array_normalizado = array_imagem.astype(float) / 255.0
    
    # Cria a matriz 3D
    matriz_3d = np.zeros((profundidade, *array_normalizado.shape))
    for i in range(profundidade):
        matriz_3d[i] = array_normalizado * (1 - i/profundidade)
    
    return matriz_3d

def visualizar_raio_x_3d(imagem, profundidade=20, surface_count=20):
    """Cria uma visualização 3D do Raio-X a partir da imagem."""
    # Converte a imagem para uma matriz 3D
    matriz_3d = converter_raio_x_para_3d(imagem, profundidade)
    
    # Reduz a resolução da matriz se necessário
    if matriz_3d.size > 5_000_000:  # Ajuste este valor conforme necessário
        fator = (5_000_000 / matriz_3d.size) ** (1/3)
        matriz_3d = reduzir_resolucao_matriz(matriz_3d, fator=fator)
    
    z, y, x = matriz_3d.shape
    
    # Cria uma malha de coordenadas
    X, Y, Z = np.mgrid[0:x:complex(0, x), 0:y:complex(0, y), 0:z:complex(0, z)]
    
    # Cria figura 3D
    fig = go.Figure(data=go.Volume(
        x=X.flatten(),
        y=Y.flatten(),
        z=Z.flatten(),
        value=matriz_3d.flatten(),
        isomin=0.1,
        isomax=0.8,
        opacity=0.1,
        surface_count=surface_count,
        colorscale='Greys',
    ))
    
    # Configura o layout
    fig.update_layout(
        scene=dict(
            xaxis_title='X',
            yaxis_title='Y',
            zaxis_title='Z',
            aspectmode='data'
        ),
        width=700,
        height=700,
        title="Visualização 3D de Raio-X"
    )
    
    return fig

def pagina_visualizacao_3d():
    st.header("Visualização 3D de Raio-X")
    
    # Upload da imagem de Raio-X
    arquivo_carregado = st.file_uploader("Faça upload do Raio-X", type=["png", "jpg", "jpeg"], key="visualizacao_3d_uploader")
    
    if arquivo_carregado is not None:
        try:
            # Abre a imagem usando PIL
            imagem = Image.open(arquivo_carregado)
            st.image(imagem, caption="Raio-X Original", use_column_width=True)
            
            # Botão para converter a imagem em 3D
            if st.button("Converter para 3D"):
                with st.spinner("Convertendo para 3D..."):
                    # Gera o gráfico 3D
                    fig_3d = visualizar_raio_x_3d(imagem, profundidade=20, surface_count=20)
                    st.plotly_chart(fig_3d, use_container_width=True)
                    st.success("Visualização 3D gerada com sucesso!")
        
        except Exception as e:
            st.error(f"Erro ao processar a imagem: {str(e)}")
    else:
        st.info("Por favor, faça o upload de uma imagem de Raio-X.")

def main():
    st.title("MedVision")
    
    try:
        inicializar_arquivo_login()  # Inicializa o arquivo de login
        
        if not st.session_state.logado:
            pagina_login()  # Exibe a página de login se não estiver logado
        else:
            st.sidebar.title(f"Bem-vindo, {st.session_state.nome_usuario}")
            if st.sidebar.button("Sair"):
                st.session_state.logado = False
                st.session_state.nome_usuario = None
                st.session_state.setores_usuario = []
                st.experimental_rerun()  # Reinicia a interface ao deslogar

            # Definindo opções do menu
            opcoes = ["Classificar Exame", "Visualizar Histórico do Paciente", "Comparar Pacientes", "Visualização 3D de Raio-X"]
            if st.session_state.nome_usuario == 'admin':
                opcoes.append("Gerenciamento de Usuários")

            opcao_menu = st.sidebar.radio("Escolha uma opção:", opcoes)

            # Chamando as funções correspondentes às opções selecionadas
            if opcao_menu == "Classificar Exame":
                st.header("Classificar Exame")
                id_paciente = st.text_input("ID do Paciente")
                opcao_modelo = st.selectbox("Escolha o modelo", [f"{setor}_{modelo}" for setor in caminhos_modelos for modelo in caminhos_modelos[setor]])
                arquivo_carregado = st.file_uploader("Faça upload da imagem de exame", type=["png", "jpg", "jpeg"])
                if st.button("Classificar Exame"):
                    classificar_exame(id_paciente, opcao_modelo, arquivo_carregado)
            
            elif opcao_menu == "Visualizar Histórico do Paciente":
                st.header("Histórico do Paciente")
                id_paciente = st.text_input("ID do Paciente para visualização do histórico")
                if st.button("Visualizar Histórico"):
                    visualizar_historico_paciente(id_paciente)
            
            elif opcao_menu == "Comparar Pacientes":
                st.header("Comparar Pacientes")
                comparar_pacientes()
            
            elif opcao_menu == "Visualização 3D de Raio-X":
                pagina_visualizacao_3d()
            
            elif opcao_menu == "Gerenciamento de Usuários" and st.session_state.nome_usuario == 'admin':
                gerenciar_usuarios()

    except Exception as e:
        st.error(f"Ocorreu um erro inesperado: {str(e)}")

if __name__ == "__main__":
    main()
